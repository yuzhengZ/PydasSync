#!/usr/bin/python
# -*- coding: utf-8 -*-

################################################################################
#
#
# Copyright 2014 Kitware Inc. 28 Corporate Drive,
# Clifton Park, NY, 12065, USA.
#
# All rights reserved.
#
# Licensed under the Apache License, Version 2.0 ( the "License" );
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
################################################################################

"""
setMataData: A helper script to set metadata by reading an excel file.
"""

import os
import sys
import getopt
import shutil
import datetime
import pprint
import pydas
import openpyxl
        
class MidasSetting(object):
    """
    Class for Pydas login setting
    """
    def __init__(self, midas_url, midas_apikey, 
                 midas_user_email, midas_root_folder_id):
        self.midas_url = midas_url
        self.midas_apikey = midas_apikey
        self.midas_user_email = midas_user_email
        self.midas_root_folder_id = midas_root_folder_id

   
class ExcelSetting(object):
    """
    Class for excel source file setting setting
    """
    def __init__(self, excel_file, excel_sheet_name,
                 excel_max_columns, excel_max_rows):
        self.excel_file = excel_file
        self.excel_sheet_name = excel_sheet_name
        self.excel_max_columns = excel_max_columns
        self.excel_max_rows = excel_max_rows

class Usage(Exception):
    def __init__(self, msg):
        self.msg = msg


def _get_midas_resource_ancestor(midas_resource_id, type='folder', root_folder_id=None):
    """
    Helper function to get the ancestor list for a Midas folder
    """
    ancestor_list = [ ]
    if (type == 'item'):
        item_info = pydas.session.communicator.item_get(
            pydas.session.token, midas_resource_id)
        ancestor_list.append(item_info['name'])
        current_folder_id = item_info['folder_id']
    else:
        current_folder_id = midas_resource_id
    while int(current_folder_id) > 0:
        if current_folder_id == root_folder_id:
            return 'folder', ancestor_list
        folder_info = pydas.session.communicator.folder_get(
            pydas.session.token, current_folder_id)
        ancestor_list.append(folder_info['name'])
        current_folder_id = folder_info['parent_id']
    id = ancestor_list[-1].split('_', 1)[-1]
    ancestor_list[-1] = id
    if current_folder_id == '-2':
        return 'community',  ancestor_list
    elif current_folder_id == '-1':
        return 'user', ancestor_list


def _midas_permision_check(midas_folder_id):
    """
    Check if the user is allowed to set matadata to items in a Midas folder
    Permission rules: 
        1) a Midas user is allowed to set matadata to items in its own folders
        2) TODO: a Midas user is allowed to set metadata to items in its community folders 
    """
    pydas_user_info = pydas.session.communicator.get_user_by_email(pydas.session.email)
    ancestor_type, ancestor_list = _get_midas_resource_ancestor(midas_folder_id)
    id = ancestor_list[-1]
    # check if it is the user himself
    if ancestor_type == 'user' and id == pydas_user_info['user_id']:
        return True
    # check if the user joins the community
    elif ancestor_type == 'community':
        community_info = pydas.session.communicator.get_community_by_id(
            id, pydas.session.token)
        community_member_group_id = community_info['membergroup_id']
        # TODO: check if the user joins the community via Pydas
        return True
    else:
        return False


def _query_yes_no(question, default="no"):
    """Helper function to ask a yes/no question via raw_input() and return their answer.
    "question" is a string that is presented to the user.
    "default" is the presumed answer if the user just hits <Enter>.
        It must be "yes" (the default), "no" or None (meaning
        an answer is required of the user).
    The "answer" return value is one of "yes" or "no".
    """
    valid = {"yes":True,   "y":True,  "ye":True,
             "no":False,     "n":False}
    if default == None:
        prompt = " [y/n] "
    elif default == "yes":
        prompt = " [Y/n] "
    elif default == "no":
        prompt = " [y/N] "
    else:
        raise ValueError("invalid default answer: '%s'" % default)

    while True:
        sys.stdout.write(question + prompt)
        choice = raw_input().lower()
        if default is not None and choice == '':
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Please respond with 'yes' or 'no' "\
                             "(or 'y' or 'n').\n")    


def sanity_check(midas_setting, excel_setting):
    """
    Sanity check for input parameters
    """
    if not os.path.isfile(excel_setting.excel_file):
        print ("Caught a sanity check error: Metadata source file %s does not exist!" \
              % excel_setting.excel_file)
        return False
    elif os.path.splitext(excel_setting.excel_file)[1] not in ['.xls', '.xlsx']:
        print ("Caught a sanity check error: Metadata source file %s is not an excel file!" \
                % excel_setting.excel_file)
        return False      
    try:      
        pydas.login(email=midas_setting.midas_user_email, 
            api_key=midas_setting.midas_apikey, url=midas_setting.midas_url)
        pydas.session.communicator.folder_get(pydas.session.token, 
                                              midas_setting.midas_root_folder_id)
        return _midas_permision_check(midas_setting.midas_root_folder_id)
    except pydas.exceptions.PydasException as detail:
        print "Caught PydasException: ", detail
        return False
    return True


def _how_to_read_metadata(metadata_names):
    """
    Define the data structure in the given excel file so that the script can 
    understand it. 
    """
    # Assumptions:
    # 1) One and only one column named "Scan No." must exist
    # 2) At least one column named "Age @ Scan" must exist
    # 3) Each "Age @ scan" column defines a metadata group. The columns between
    # it and the next "Age @ scan" column only belong to this metadata group
    # 4) All columns before the first "Age @ Scan" column are common metadata
    # and belong to all the metadata group.
    # 5) Empty columns are not allowed (except for undefined ones at the end)
    scan_number_col_id = None
    age_at_scan_col_ids = []
    for idx, name in enumerate(metadata_names):
        if (name == 'Scan No.'):
    # Note: excel uses 1-based index and python list use 0-based index
            scan_number_col_id = idx + 1
        elif (name == "Age @ Scan"):
            age_at_scan_col_ids.append(idx + 1)
    if ((scan_number_col_id is None) or (not age_at_scan_col_ids)):
        print ("Caught a sanity check error: 'Scan No.' and/or 'Age @ Scan' \
            column do not exist!")
        return False
    metadata_structure = {}
    metadata_structure['scan_number_col_id'] = scan_number_col_id
    metadata_structure['age_at_scan_col_ids'] = age_at_scan_col_ids
    metadata_structure['metadata_groups'] = {}
    #  metadata_structure['metadata_groups'] is defined as below
    #  scan_number_col_id -> age_at_group_id -> [column_ids]
    # E.g. metadata_structure['metadata_groups'] =
    #         { 4: [1, 2, 3, 4, 5, 6],
    #           7: [1, 2, 3, 7, 8, 9],
    #          10: [1, 2, 3, 10, 11, 12],
    #          13: [1, 2, 3, 13, 14, 15],
    #          16: [1, 2, 3, 16, 17, 18]}
    # Hack, manually set data structure for now
    metadata_structure['metadata_groups'][4] =  [1, 2, 3, 4, 5, 6]
    metadata_structure['metadata_groups'][7] =  [1, 2, 3, 7, 8, 9]
    metadata_structure['metadata_groups'][10] = [1, 2, 3, 10, 11, 12]
    metadata_structure['metadata_groups'][13] = [1, 2, 3, 13, 14, 15]
    metadata_structure['metadata_groups'][16] = [1, 2, 3, 16, 17, 18]
    return metadata_structure

def _get_metadata_from_excel(excel_setting):
    """
    Read an excel file to get all the metadata.
    """
    print "\nStart parsing metadata from the sheet %s in the excel file %s." % (
        excel_setting.excel_sheet_name, excel_setting.excel_file)
    wb = openpyxl.load_workbook(filename = excel_setting.excel_file)
    sheet = wb.active
    if (excel_setting.excel_sheet_name is not None):
        sheet = wb[excel_setting.excel_sheet_name]
    # Assumption: metadata names are defined in the first row
    metadata_names = []
    main_row = 1
    for col_idx in xrange(1, excel_setting.excel_max_columns):
         col = openpyxl.cell.get_column_letter(col_idx)
         name = sheet['%s%s'%(col, main_row)].value
         if name is not None:
            metadata_names.append(name)
    excel_structure = _how_to_read_metadata(metadata_names)
    if not excel_structure:
        return False
    metadata_dict = {}
    for row in xrange(main_row + 1, excel_setting.excel_max_rows):
        # Assumption: the row is not empty if its first column has value
        if sheet['%s%s'%('A', row)].value is None:
            continue
        scan_number_col = openpyxl.cell.get_column_letter(excel_structure['scan_number_col_id'])
        scan_number = sheet['%s%s'%(scan_number_col, row)].value
        if not scan_number in metadata_dict.keys():
            metadata_dict[scan_number] = {}
        for age_at_scan_col_id in excel_structure['age_at_scan_col_ids']:
            age_at_scan_col = openpyxl.cell.get_column_letter(age_at_scan_col_id)       
            age_at_scan = sheet['%s%s'%(age_at_scan_col, row)].value
            if not age_at_scan in metadata_dict[scan_number]:
                metadata_dict[scan_number][age_at_scan] = {}
            for col_id in excel_structure['metadata_groups'][age_at_scan_col_id]:
                col = openpyxl.cell.get_column_letter(col_id)
                # Note: excel uses 1-based index and python list use 0-based index
                value = sheet['%s%s'%(col, row)].value
                if (type(value) is datetime.date) or (type(value) is datetime.datetime):
                    value = value.strftime("%b %d, %Y")
                metadata_dict[scan_number][age_at_scan][metadata_names[col_id - 1]] = str(value)  
    # Double check the data structure
    if not metadata_dict:
        print ("We didn't get any metadata from the excel file, is it empty?")
    pp = pprint.PrettyPrinter(indent = 2)
    pp.pprint(metadata_dict.itervalues().next())
    agree_to_continue = _query_yes_no("A sample metadata group is displayed as above. "  \
            "Do you think the excel file is parsed correctly?")
    if not agree_to_continue:
        return False
    else:
        return metadata_dict                


def set_matadata(midas_setting, excel_setting):
    """
    Set metadata to the items in Midas
    """
    metadata_lookup = {}
    metadata_lookup = _get_metadata_from_excel(excel_setting)
    if not metadata_lookup:
        return False
    print "\nStart set metadata to Midas. Please be patient!"
    # upload 'local_only' data to Midas
    midas_folders_ids = []
    midas_children_items = { }
    midas_folders_ids.append(midas_setting.midas_root_folder_id)
    while midas_folders_ids:
        folder_id = midas_folders_ids.pop(0)
        for resource_type, resource_list in pydas.session.communicator.folder_children(
          pydas.session.token, folder_id).iteritems():
            if resource_type == 'folders':
                for midas_folder in resource_list:
                    midas_folders_ids.append(midas_folder['folder_id'])
            elif resource_type == 'items':
                for midas_item in resource_list:
                    midas_children_items[midas_item['item_id']] = midas_item
    metadata_info = []
    temp_list = []
    #Assumption: item is name as xxx_xx_months_xxxxx.....     
    for item_id, item_info in midas_children_items.iteritems():
         metadata_info[:] = []
         temp_list[:] = []
         temp_list  = item_info['name'].split('_')
         scan_number = int(temp_list[0])
         age_at_scan = temp_list[1]
         if (scan_number in metadata_lookup.keys() \
           and age_at_scan in metadata_lookup[scan_number].keys()):
             metadata_info = pydas.session.communicator.get_item_metadata(item_id, pydas.session.token)
             for k, v in metadata_lookup[scan_number][age_at_scan].iteritems():
                found = False
                for metadata in metadata_info:
                    if (metadata['element'] == k and metadata['value'] == v):
                        found = True
                        break
                if not found:
                    pydas.session.communicator.set_item_metadata(pydas.session.token, item_id, k, v)
             print "\nMetadata for item %s (item id is %s) is up to date." \
               % (item_info['name'], item_id)
         else:
             print "\nCannot find the metadata for item %s (item id is %s)." \
               % (item_info['name'], item_id)
    
    print ("MetaData in %s has been set to all the items in %s.\n" \
            % (excel_setting.excel_file, os.path.join(midas_setting.midas_url, 
               'folder', midas_setting.midas_root_folder_id)))
    return True
     
def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], "hx:s:u:e:a:f:", 
            ["help", "excelfile=", "sheetname=", "url=", "email=", "apikey=", "folderid=" ])
    except getopt.error, msg:
        raise Usage(msg)

    # default setting
    excel_file = None
    excel_sheet_name = None
    excel_max_columns = 30
    excel_max_rows = 30
    midas_url = None
    midas_user_email = None
    midas_apikey = None
    midas_root_folder_id = None

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            print "setMetaData.py -x <metadata_excel_file_path> [-s <excel_sheet_name>]" \
                "-u <midas_url> -e <midas_user_email>  -a <midas_api_key> -f <midas_folder_id>"
            sys.exit()
        elif opt in ("-x", "--excelfile"):
            excel_file = arg
        elif opt in ("-s", "--sheetname"):
            excel_sheet_name = arg
        elif opt in ("-u", "--url"):
            midas_url = arg.rstrip('/')
        elif opt in ("-e", "--email"):
            midas_user_email = arg
        elif opt in ("-a", "--apikey"):
            midas_apikey = arg
        elif opt in ("-f", "--folderid"):
            midas_root_folder_id = arg

    # sanity check for input parameters
    for param in [excel_file, midas_url, midas_user_email, midas_apikey, 
                  midas_root_folder_id]:
        if param is None:
            print "Caught a sanity check error: At least one required parameter is missing!"
            print "setMetadata.py --help for more information"
            sys.exit()
    midas_url = midas_url.rstrip('/')
    excel_file = os.path.abspath(excel_file)
    midas_setting = MidasSetting(midas_url, midas_apikey, 
        midas_user_email, midas_root_folder_id)
    excel_setting = ExcelSetting(excel_file, excel_sheet_name,
        excel_max_columns, excel_max_rows)
    input_sanity = sanity_check(midas_setting, excel_setting)
    
    # set matadata
    if input_sanity:
        set_matadata(midas_setting, excel_setting)
            

if __name__ == "__main__":
    sys.exit(main())
